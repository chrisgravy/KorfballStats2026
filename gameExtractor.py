import json
from openpyxl import load_workbook
from datetime import datetime
import re

wb = load_workbook("2026-KSA-State-League-Timetable.xlsx", data_only=True)
ws = wb["2026 timetable"]

rows = list(ws.iter_rows(values_only=True))

TEAM_MAP = {
    "Boomers": "Adelaide Boomers",

    "Scaldis": "Scaldis Tigers",

    "North": "North Adelaide Roosters",

    "Arista": "Arista Marion",

    "Flinders": "Flinders Sharks",
}

matches = []

for i, row in enumerate(rows):

    # -------------------------
    # Round Row
    # -------------------------
    if row[1] and isinstance(row[1], str) and row[1].startswith("Round"):

        current_round = row[1]

        venue_row = rows[i + 1]
        current_venue = venue_row[1]

        match = re.search(
            r"Round\s+(\d+)\s+-\s+(\d+/\d+/\d+)",
            row[1]
        )

        if match:
            round_number = match.group(1)

            current_date = datetime.strptime(
                match.group(2),
                "%d/%m/%y"
            ).date()


    # -------------------------
    # Division 1
    # -------------------------
    division = row[9]

    if division == 1:

        match_time = row[3]
        home_team = row[10]
        away_team = row[11]

        if match_time and home_team and away_team:

            start_datetime = datetime.combine(
                current_date,
                match_time
            )

            matches.append({
                "round": current_round,
                "division": division,
                "date": current_date.isoformat(),
                "datetime": start_datetime.isoformat(),
                "venue": current_venue,

                "home_team": home_team,
                "away_team": away_team,

                "home_club": TEAM_MAP.get(home_team.strip(), home_team),
                "away_club": TEAM_MAP.get(away_team.strip(), away_team)
            })

    # -------------------------
    # Division 2
    # -------------------------
    division = row[4]

    if division == 2:

        match_time = row[3]
        home_team = row[5]
        away_team = row[6]
        if match_time and home_team and away_team:

            start_datetime = datetime.combine(
                current_date,
                match_time
            )

            matches.append({
                "round": current_round,
                "division": division,
                "date": current_date.isoformat(),
                "datetime": start_datetime.isoformat(),
                "venue": current_venue,

                "home_team": home_team,
                "away_team": away_team,

                "home_club": TEAM_MAP.get(home_team.strip(), home_team),
                "away_club": TEAM_MAP.get(away_team.strip(), away_team)
            })

print(f"Found {len(matches)} matches")


with open("matches.json", "w") as f:
    json.dump(matches, f, indent=4)
